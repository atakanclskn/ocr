# Gerekli kütüphaneleri import et
# Not: Gemini API için önce kurulum yapın: pip install google-generativeai openpyxl
import requests
import time
import json
import os
from datetime import datetime
try:
    import google.generativeai as genai
    GEMINI_AVAILABLE = True
except ImportError:
    GEMINI_AVAILABLE = False
    print("⚠️  Gemini API kütüphanesi kurulu değil!")
    print("   Kurmak için: pip install google-generativeai")
    print("   OCR işlemi devam edecek ancak Gemini analizi yapılamayacak.\n")

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️  OpenPyXL kütüphanesi kurulu değil!")
    print("   Kurmak için: pip install openpyxl")
    print("   Excel çıktısı oluşturulamayacak.\n")

API_KEY = "lHHnyClmPHpxAPMCLDqDtMykU8U2kON7lLG9TOuRVNtV4cHxVtCOTaxIXjkCiBQE"
GEMINI_API_KEY = "AIzaSyBnPw6fjlSk4BDDLAG-IbQItB5N03ZmqPs"  # Gemini API anahtarınızı buraya girin
UPLOAD_URL = "https://backend.scandocflow.com/v1/api/documents/extractAsync"
STATUS_URL = "https://backend.scandocflow.com/v1/api/documents/status"
file_path = "./image8.jpg"
excel_template_path = "./beyanname şablon.xlsx"  # Excel şablon dosyası

# Gemini API yapılandırması
if GEMINI_AVAILABLE and GEMINI_API_KEY != "YOUR_GEMINI_API_KEY":
    genai.configure(api_key=GEMINI_API_KEY)
elif GEMINI_API_KEY == "YOUR_GEMINI_API_KEY":
    print("⚠️  Gemini API anahtarı ayarlanmamış! Kod içinde GEMINI_API_KEY değişkenini güncelleyin.\n")

# OCR işlemi için parametreler
payload = {
    "type": "ocr",
    "lang": "tur",
    "retain": "true",
}

# Dosya yükleyip OCR işlemini başlat
print("Dosya yükleniyor...")
with open(file_path, "rb") as file:
    response = requests.post(
        UPLOAD_URL,
        params={"access_token": API_KEY},
        data=payload,
        files={"files": file}
    )

print(f"Upload yanıt kodu: {response.status_code}")
print(f"Upload yanıtı: {response.text}")

if response.status_code == 200:
    result = response.json()
    request_id = result.get("id")
    
    if not request_id:
        print("Request ID bulunamadı. Tam yanıt:")
        print(json.dumps(result, indent=2, ensure_ascii=False))
        exit()
    
    print(f"İşlem başladı, Request ID: {request_id}")
    print(f"Status: {result.get('status')}")
    
    if result.get("webhook"):
        print(f"Webhook durumu: {result['webhook'].get('status')}")
else:
    print(f"Yükleme sırasında hata: {response.status_code}, {response.text}")
    exit()

# İşlem tamamlanana kadar status kontrolü yap
print("\nOCR işlemi devam ediyor, status kontrol ediliyor...")

max_wait_time = 120  # 2 dakika (dokümantasyonda 2 dakikaya kadar sürebileceği belirtilmiş)
time_waited = 0
interval = 5
success = False

while time_waited < max_wait_time:
    print(f"\n[{time_waited}s] Status kontrol ediliyor...")
    
    # Status kontrolü için doğru endpoint ve parametreler
    status_params = {
        "access_token": API_KEY,
        "request_id": request_id
    }
    
    response = requests.get(STATUS_URL, params=status_params)
    
    print(f"Status yanıt kodu: {response.status_code}")
    
    if response.status_code == 200:
        result = response.json()
        status = result.get("status", "").lower()
        
        print(f"İşlem durumu: {status}")
        
        if status == "success" or status == "completed":
            print("\nOCR işlemi başarıyla tamamlandı!")
            
            # Webhook durumunu kontrol et
            if result.get("webhook"):
                webhook_status = result["webhook"].get("status")
                print(f"Webhook durumu: {webhook_status}")
            
            # Dokümanları işle
            documents = result.get("documents", [])
            
            if documents:
                print(f"\n{len(documents)} doküman bulundu.")
                
                all_text = ""
                
                for idx, doc in enumerate(documents):
                    print(f"\n--- Doküman {idx + 1} ---")
                    print(f"Tip: {doc.get('type')}")
                    
                    # textAnnotation içindeki metni çıkar
                    text_annotation = doc.get("textAnnotation", {})
                    pages = text_annotation.get("Pages", [])
                    
                    doc_text = ""
                    
                    for page_idx, page in enumerate(pages):
                        print(f"Sayfa {page_idx + 1}: {len(page.get('Words', []))} kelime bulundu")
                        
                        words = page.get("Words", [])
                        
                        # Kelimeleri birleştir
                        page_text = " ".join([word.get("Text", "") for word in words])
                        doc_text += page_text + "\n\n"
                    
                    all_text += f"=== Doküman {idx + 1} ===\n{doc_text}\n\n"
                
                if all_text.strip():
                    print(f"\nToplam {len(all_text)} karakter metin çıkarıldı.")
                    
                    # Dosya adını ve tarih bilgisini hazırla
                    base_filename = os.path.basename(file_path)
                    name_without_ext = os.path.splitext(base_filename)[0]
                    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                    output_filename = f"{name_without_ext}_{timestamp}.txt"
                    
                    # Metni dosyaya kaydet
                    with open(output_filename, "w", encoding="utf-8") as f:
                        f.write(all_text)
                    print(f"✅ Metin '{output_filename}' dosyasına kaydedildi.")
                    
                    # İlk 500 karakteri göster
                    print("\n--- Çıkarılan Metin (İlk 500 karakter) ---")
                    print(all_text[:500] + "..." if len(all_text) > 500 else all_text)
                    print("--- Metin Sonu ---")
                    
                    # OCR sonucunu Gemini'ye gönder
                    if GEMINI_AVAILABLE and GEMINI_API_KEY != "YOUR_GEMINI_API_KEY":
                        print("\n\n📤 Beyanname bilgileri çıkarılıyor...")
                        
                        try:
                            # Gemini model oluştur (güncel model adını kullan)
                            model = genai.GenerativeModel('gemini-2.5-flash')  # veya 'gemini-1.5-pro'
                            
                            # Prompt hazırla
                            prompt = f"""
                            Sen bir gümrük beyannamesi uzmanısın. Aşağıdaki OCR ile çıkarılmış beyanname metnini analiz et ve sadece istenen bilgileri çıkar.
                            
                            İstenen bilgiler ve açıklamaları:
                            - Alıcı: İthalatçı firma adı
                            - ALICI VKN: Alıcı firma vergi kimlik numarası (10 veya 11 haneli sayı)
                            - KONTEYNER NO: Konteyner numarası (varsa)
                            - Teslim şekli: Teslim şekli bilgisi (örn: EXW, FCA, FOB, CIF vb.)
                            - Brüt KG: Brüt ağırlık kilogram cinsinden
                            - SON AMBAR: Gümrük müdürlüğü adı
                            - ÖZET BEYAN NO: Beyanname numarası
                            - BEYANNAME TESCİL TARİHİ: Beyanname tescil tarihi
                            - TAREKS-TARIM-TSE: Belgelerde TAREKS, TARIM veya TSE ibaresi geçiyorsa VAR, geçmiyorsa YOK yaz
                            
                            ÇIKTI FORMATI (SADECE BU FORMATTA YAZ, BAŞKA BİR ŞEY EKLEME):
                            Alıcı: [değer]
                            ALICI VKN: [değer]
                            KONTEYNER NO: [değer veya "Belirtilmemiş"]
                            Teslim şekli: [değer veya "Belirtilmemiş"]
                            Brüt KG: [değer veya "Belirtilmemiş"]
                            SON AMBAR: [değer]
                            ÖZET BEYAN NO: [değer]
                            BEYANNAME TESCİL TARİHİ: [değer]
                            TAREKS-TARIM-TSE: [VAR veya YOK]
                            
                            OCR Metni:
                            {all_text}
                            """
                            
                            # Gemini'ye gönder
                            response = model.generate_content(prompt)
                            
                            # Gemini yanıtını göster
                            print("\n✅ Gemini API yanıtı alındı!")
                            print("\n" + "="*50)
                            print("BEYANNAME BİLGİLERİ (EXCEL FORMATI):")
                            print("="*50)
                            print(response.text)
                            print("="*50)
                            
                            # Gemini yanıtını da dosyaya kaydet
                            gemini_filename = f"{name_without_ext}_{timestamp}_beyanname_bilgileri.txt"
                            with open(gemini_filename, "w", encoding="utf-8") as f:
                                f.write("BEYANNAME BİLGİLERİ (EXCEL FORMATI)\n")
                                f.write("="*50 + "\n\n")
                                f.write(response.text)
                                f.write("\n\n" + "="*50 + "\n")
                                f.write("\nORİJİNAL OCR METNİ:\n")
                                f.write("="*50 + "\n")
                                f.write(all_text)
                            
                            print(f"\n📁 Beyanname bilgileri '{gemini_filename}' dosyasına kaydedildi.")
                            
                            # Excel şablonuna verileri yaz
                            if EXCEL_AVAILABLE:
                                try:
                                    # Gemini yanıtını parse et
                                    lines = response.text.strip().split('\n')
                                    data_dict = {}
                                    for line in lines:
                                        if ':' in line:
                                            key, value = line.split(':', 1)
                                            data_dict[key.strip()] = value.strip()
                                    
                                    # Şablon Excel dosyasını aç
                                    if os.path.exists(excel_template_path):
                                        workbook = openpyxl.load_workbook(excel_template_path)
                                        sheet = workbook.active
                                        
                                        # Sütun başlıklarını bul
                                        column_mapping = {}
                                        for col in range(1, sheet.max_column + 1):
                                            header = sheet.cell(row=1, column=col).value
                                            if header:
                                                header = header.strip()
                                                if header == "Alıcı D.Ö":
                                                    column_mapping["Alıcı"] = col
                                                elif header == "ALICI VKN":
                                                    column_mapping["ALICI VKN"] = col
                                                elif header == "KONTEYNER NO":
                                                    column_mapping["KONTEYNER NO"] = col
                                                elif header == "Teslim şekli":
                                                    column_mapping["Teslim şekli"] = col
                                                elif header == "Brüt KG":
                                                    column_mapping["Brüt KG"] = col
                                                elif header == "SON AMBAR":
                                                    column_mapping["SON AMBAR"] = col
                                                elif header == "ÖZET BEYAN NO":
                                                    column_mapping["ÖZET BEYAN NO"] = col
                                                elif header == "BEYANNAME TESCİL TARİHİ":
                                                    column_mapping["BEYANNAME TESCİL TARİHİ"] = col
                                                elif header == "TAREKS-TARIM-TSE (VAR-YOK)":
                                                    column_mapping["TAREKS-TARIM-TSE"] = col
                                        
                                        # İlk boş satırı bul
                                        next_row = 2
                                        for row in range(2, sheet.max_row + 2):
                                            if not any(sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)):
                                                next_row = row
                                                break
                                        
                                        # Verileri yaz
                                        for field, col in column_mapping.items():
                                            if field in data_dict:
                                                sheet.cell(row=next_row, column=col).value = data_dict[field]
                                        
                                        # Excel dosyasını kaydet
                                        excel_filename = f"{name_without_ext}_{timestamp}_beyanname_dolu.xlsx"
                                        workbook.save(excel_filename)
                                        workbook.close()
                                        
                                        print(f"\n✅ Excel şablonu dolduruldu: '{excel_filename}'")
                                        print(f"   Veriler {next_row}. satıra eklendi.")
                                        
                                    else:
                                        print(f"\n⚠️  Excel şablonu '{excel_template_path}' bulunamadı!")
                                        
                                        # Yeni Excel oluştur
                                        workbook = openpyxl.Workbook()
                                        sheet = workbook.active
                                        
                                        # Başlıkları ekle
                                        headers = ["Alıcı D.Ö", "ALICI VKN", "Nakliyeci", "Fat. Tarihi", 
                                                  "Tahmini Çıkış Tarihi", "Varış Tarihi", "Çıkış Limanı", 
                                                  "HBL", "KONTEYNER NO", "Teslim şekli", "Brüt KG", 
                                                  "Hacim", "Rakip Navlun w/m", "Navlun Fatura Tutarı", 
                                                  "Rakip EXW / FCA All in Fatura Tutarı", "Konsol/Komple", 
                                                  "Öykü Dönem", "Navlun w/m Total", "Fark w/m", "Varış Limanı", 
                                                  "SON AMBAR", "HAT", "ÖZET BEYAN NO", "BEYANNAME TESCİL TARİHİ", 
                                                  "TAREKS-TARIM-TSE (VAR-YOK)", "KAYIT TARİHİ"]
                                        
                                        for col, header in enumerate(headers, 1):
                                            sheet.cell(row=1, column=col).value = header
                                            sheet.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
                                        
                                        # Verileri 2. satıra ekle
                                        sheet.cell(row=2, column=1).value = data_dict.get("Alıcı", "")
                                        sheet.cell(row=2, column=2).value = data_dict.get("ALICI VKN", "")
                                        sheet.cell(row=2, column=9).value = data_dict.get("KONTEYNER NO", "")
                                        sheet.cell(row=2, column=10).value = data_dict.get("Teslim şekli", "")
                                        sheet.cell(row=2, column=11).value = data_dict.get("Brüt KG", "")
                                        sheet.cell(row=2, column=21).value = data_dict.get("SON AMBAR", "")
                                        sheet.cell(row=2, column=23).value = data_dict.get("ÖZET BEYAN NO", "")
                                        sheet.cell(row=2, column=24).value = data_dict.get("BEYANNAME TESCİL TARİHİ", "")
                                        sheet.cell(row=2, column=25).value = data_dict.get("TAREKS-TARIM-TSE", "")
                                        sheet.cell(row=2, column=26).value = datetime.now().strftime("%d.%m.%Y")
                                        
                                        # Sütun genişliklerini ayarla
                                        for col in range(1, 27):
                                            sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
                                        
                                        excel_filename = f"{name_without_ext}_{timestamp}_beyanname_yeni.xlsx"
                                        workbook.save(excel_filename)
                                        workbook.close()
                                        
                                        print(f"\n✅ Yeni Excel dosyası oluşturuldu: '{excel_filename}'")
                                        
                                except Exception as e:
                                    print(f"\n❌ Excel işleme hatası: {e}")
                            
                            # CSV formatında da kaydet (eski kod korunuyor)
                            try:
                                # CSV dosyası oluştur
                                csv_filename = f"{name_without_ext}_{timestamp}_beyanname.csv"
                                with open(csv_filename, "w", encoding="utf-8-sig") as f:
                                    # Başlıkları yaz
                                    headers = ["Alıcı", "ALICI VKN", "KONTEYNER NO", "Teslim şekli", 
                                              "Brüt KG", "SON AMBAR", "ÖZET BEYAN NO", 
                                              "BEYANNAME TESCİL TARİHİ", "TAREKS-TARIM-TSE"]
                                    f.write(";".join(headers) + "\n")
                                    
                                    # Değerleri yaz
                                    values = []
                                    for header in headers:
                                        values.append(data_dict.get(header, ""))
                                    f.write(";".join(values) + "\n")
                                
                                print(f"\n📊 CSV dosyası '{csv_filename}' olarak da kaydedildi.")
                                
                            except Exception as e:
                                print(f"CSV oluşturma hatası: {e}")
                            
                        except Exception as e:
                            print(f"\n❌ Gemini API hatası: {e}")
                            print("Not: Gemini API anahtarınızı ve model adını kontrol edin.")
                            print("Mevcut modeller: gemini-1.5-flash, gemini-1.5-pro")
                    
                    # API'den doküman ID'sini al ve web export linkini oluştur
                    if documents:
                        first_doc = documents[0]
                        if "id" in first_doc:
                            doc_id = first_doc["id"]
                            print(f"\n📄 Doküman ID: {doc_id}")
                            print(f"🌐 Web arayüzünde görüntülemek için:")
                            print(f"   https://app.scandocflow.com/documents/{doc_id}")
                    
                    success = True
                else:
                    print("\nUYARI: Doküman bulundu ancak metin çıkarılamadı!")
                    
                    # Debug için tam yanıtı kaydet
                    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                    debug_filename = f"api_response_{timestamp}.json"
                    with open(debug_filename, "w", encoding="utf-8") as f:
                        json.dump(result, f, indent=2, ensure_ascii=False)
                    print(f"Tam API yanıtı '{debug_filename}' dosyasına kaydedildi.")
            else:
                print("\nUYARI: Hiç doküman bulunamadı!")
                print("Tam yanıt:")
                print(json.dumps(result, indent=2, ensure_ascii=False))
            
            break
            
        elif status == "running" or status == "processing":
            print("İşlem devam ediyor...")
            
            # İşlem yüzdesini göster (varsa)
            if "progress" in result:
                print(f"İlerleme: %{result.get('progress', 0)}")
                
        elif status == "failed" or status == "error":
            print(f"\nİşlem başarısız oldu!")
            
            if "error" in result:
                print(f"Hata mesajı: {result.get('error')}")
            
            # Tam yanıtı göster
            print("Tam yanıt:")
            print(json.dumps(result, indent=2, ensure_ascii=False))
            break
            
        else:
            print(f"Bilinmeyen durum: {status}")
            print("Tam yanıt:")
            print(json.dumps(result, indent=2, ensure_ascii=False))
            
    else:
        print(f"Status kontrolü başarısız: {response.status_code}")
        
        try:
            error_detail = response.json()
            print(f"Hata detayı: {json.dumps(error_detail, indent=2, ensure_ascii=False)}")
        except:
            print(f"Hata metni: {response.text}")
        
        # 404 hatası alıyorsak request_id yanlış olabilir
        if response.status_code == 404:
            print("Request ID bulunamadı. İşlem silinmiş veya ID yanlış olabilir.")
            break
    
    time.sleep(interval)
    time_waited += interval

if not success:
    print(f"\n\nİşlem {max_wait_time} saniye içinde tamamlanamadı.")
    print("\nKontrol edilecekler:")
    print(f"1. Request ID: {request_id}")
    print("2. API anahtarınızın geçerli olduğundan emin olun")
    print("3. Web arayüzünde işlemin durumunu kontrol edin")
    print("4. Dosya boyutunun 50MB'ı aşmadığından emin olun")
    
print("\n\nİşlem tamamlandı.")